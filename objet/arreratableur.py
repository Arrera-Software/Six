from openpyxl import Workbook , load_workbook

class CArreraTableur:
    def __init__(self,file):
        self.__file = file
        try:
            self.__workbook = load_workbook(file)
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save(self.__file)
            self.__workbook = load_workbook(file)
        self.__table = self.__workbook.active
    
    def __del__(self):
        self.__workbook.close()

    def read(self):
        sortie = {}
        for row_idx, row in enumerate(self.__table.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                cell_position = f"{self.__table.cell(row=row_idx, column=col_idx).coordinate}"
                if (value != None ):
                    sortie[cell_position] = value
        return sortie
    
    def write(self,case:str,valeur):
        try:
            self.__table[case]=valeur
            return True
        except Exception as e:
            # print(f"Erreur lors de l'écriture dans la cellule {case} : {e}")
            return False
    
    def deleteValeur(self,case):
        try :
            self.__table[case] = None
            return True
        except Exception as e:
            print(e)
            return False
    
    def saveFile(self):
        try :
            self.__workbook.save(self.__file)
            return True
        except Exception as e:
            # print(f"Erreur lors de la sauvegarde du fichier : {e}")
            return False
    
    def closeFile(self):
        self.__del__()

    def somme(self,caseDestination:str,case1:str,case2:str):
        try :
            self.__table[caseDestination] = "=SUM("+case1+":"+case2+")"
            return True
        except Exception as e:
            # print(f"Erreur lors de la somme : {e}")
            return False

    def moyenne(self,caseDestination:str,case1:str,case2:str):
        try:
            self.__table[caseDestination] = "=AVERAGE("+case1+":"+case2+")"
            return True
        except Exception as e:
            # print(f"Erreur lors de la somme : {e}")
            return False
    
    def comptage(self,caseDestination:str,case1:str,case2:str):
        try:
            self.__table[caseDestination] = "=COUNT("+case1+":"+case2+")"
            return True
        except Exception as e:
            # print(f"Erreur lors de la somme : {e}")
            return False
    
    def minimun(self,caseDestination:str,case1:str,case2:str):
        try:
            self.__table[caseDestination] = "=MIN("+case1+":"+case2+")"
            return True
        except Exception as e:
            # print(f"Erreur lors de la somme : {e}")
            return False

    def maximun(self,caseDestination:str,case1:str,case2:str):
        try:
            self.__table[caseDestination] = "=MAX("+case1+":"+case2+")"
            return True
        except Exception as e:
            # print(f"Erreur lors de la somme : {e}")
            return False